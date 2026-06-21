"""Kaynak verilerden okuma modülü.
- Rapor pptx: Hedef/Gerçekleşen metin parse + KPI OCR
- Furkan sayfası: tarih aralığına göre LFL büyüme kategorileri
"""
import re, warnings, zipfile
import numpy as np
import openpyxl
from PIL import Image
from pptx import Presentation
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# Tarih → hafta numarası eşlemesi (Mavi perakende takvimi, yaklaşık)
# ─────────────────────────────────────────────────────────────────────────────
# Haftalar Pazartesi başlar. 2026 yılı perakende hafta numaraları:
_HAFTA_TAKVIM = {
    # (ay_adi, gun_basi, gun_sonu): hafta_no
    # Haziran 2026
    ('Haziran', 1,  7):  18,
    ('Haziran', 8, 14):  19,
    ('Haziran', 15, 21): 20,
    ('Haziran', 22, 28): 21,
    ('Haziran', 29, 30): 22,
    # Mayıs 2026
    ('Mayıs', 1,  7):  14,
    ('Mayıs', 8, 14):  15,
    ('Mayıs', 15, 21): 16,
    ('Mayıs', 22, 28): 17,
    ('Mayıs', 29, 31): 18,
    # Temmuz 2026
    ('Temmuz', 1, 7):   23,
    ('Temmuz', 8, 14):  24,
    ('Temmuz', 15, 21): 25,
    ('Temmuz', 22, 28): 26,
}

_AY_TR = {'Ocak':1,'Şubat':2,'Mart':3,'Nisan':4,'Mayıs':5,'Haziran':6,
           'Temmuz':7,'Ağustos':8,'Eylül':9,'Ekim':10,'Kasım':11,'Aralık':12}

def _date_range_to_haftalar(date_range_str):
    """
    '1-17 Haziran' → son günün haftasını bulur, o haftaya kadar kümülatif.
    Döner: (ay_adi, son_hafta_no) ya da (None, None).
    """
    m = re.search(r'(\d+)\s*[-–]\s*(\d+)\s*([A-ZÇĞİÖŞÜa-züçğışö]+)', date_range_str, re.I)
    if not m:
        m2 = re.search(r'(\d+)\s*([A-ZÇĞİÖŞÜa-züçğışö]+)', date_range_str, re.I)
        if m2:
            gun_son = int(m2.group(1))
            ay_adi = m2.group(2).capitalize()
        else:
            return None, None
    else:
        gun_son = int(m.group(2))
        ay_adi = m.group(3).capitalize()

    # Ay adını normalize et
    for tr_ay in _AY_TR:
        if tr_ay.lower().startswith(ay_adi.lower()[:4]):
            ay_adi = tr_ay; break

    # Son günün haftasını bul
    son_hafta = None
    for (ay, bas, son), hf in _HAFTA_TAKVIM.items():
        if ay == ay_adi and bas <= gun_son <= son:
            son_hafta = hf; break
    # Tam eşleşme yoksa en yakın küçük haftayı al
    if son_hafta is None:
        for (ay, bas, son), hf in sorted(_HAFTA_TAKVIM.items(), key=lambda x: x[1]):
            if ay == ay_adi and bas <= gun_son:
                son_hafta = hf
    return ay_adi, son_hafta


# ─────────────────────────────────────────────────────────────────────────────
# Furkan sayfası → LFL büyüme
# ─────────────────────────────────────────────────────────────────────────────
def parse_cevirdigim(planlama_path, date_range_str=None):
    """
    Furkan sayfasından tarih aralığına göre LFL büyüme verisini çeker.
    date_range_str: '1-17 Haziran' gibi. Verilmezse en son dolu Total sütunu.
    """
    ay_adi, son_hafta = (None, None)
    if date_range_str:
        ay_adi, son_hafta = _date_range_to_haftalar(date_range_str)

    wb = openpyxl.load_workbook(planlama_path, read_only=True)
    # 'furkan' sayfasını bul
    sheet_name = None
    for sn in wb.sheetnames:
        if 'furkan' in sn.lower():
            sheet_name = sn; break
    if not sheet_name:
        # Yedek: LFL Growth sayfası
        for sn in wb.sheetnames:
            if 'LFL' in sn and 'Growth' in sn:
                sheet_name = sn; break
    wb.close()

    if not sheet_name:
        return _empty_result()

    wb2 = openpyxl.load_workbook(planlama_path, read_only=True, data_only=True)
    ws = wb2[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=110, max_col=300, values_only=True))
    wb2.close()

    ay_row  = rows[4]
    hf_row  = rows[5]
    tip_row = rows[6]

    # Hedef sütununu bul: ay + hafta='Hedef' + tip='SDG' → ALL = SDG+3
    hdf_all_col = None
    for j in range(len(tip_row)):
        ay  = str(ay_row[j])  if ay_row[j]  else ''
        hf  = str(hf_row[j])  if hf_row[j]  else ''
        tip = str(tip_row[j]) if tip_row[j] else ''
        if ay == ay_adi and hf == 'Hedef' and tip == 'SDG':
            hdf_all_col = j + 3; break
    # Yoksa genel hedef (tüm ayları gez)
    if hdf_all_col is None:
        for j in range(len(tip_row)):
            hf  = str(hf_row[j])  if hf_row[j]  else ''
            tip = str(tip_row[j]) if tip_row[j] else ''
            if hf == 'Hedef' and tip == 'SDG':
                hdf_all_col = j + 3

    # Gerçekleşen sütunu: son haftaya kadar kümülatif
    # Eğer son hafta = Total → Total ALL sütunu kullan
    # Eğer son hafta = belirli hafta → o haftanın ALL sütunu
    grc_all_col = None
    ay_sezon_col = None  # AY Sezon Hedef %

    if ay_adi and son_hafta:
        # Önce Total sütununu dene (kümülatif MTD)
        for j in range(len(tip_row)):
            ay  = str(ay_row[j])  if ay_row[j]  else ''
            hf  = str(hf_row[j])  if hf_row[j]  else ''
            tip = str(tip_row[j]) if tip_row[j] else ''
            if ay == ay_adi and hf == 'Total' and tip == 'SDG':
                all_col = j + 3
                # Dolu mu?
                for i in range(7, min(50, len(rows))):
                    v = rows[i][all_col] if all_col < len(rows[i]) else None
                    if v and v != '#REF!':
                        try:
                            float(v); grc_all_col = all_col; break
                        except: pass
                if grc_all_col: break

        # Total yoksa son hafta sütununu kullan
        if grc_all_col is None:
            for j in range(len(tip_row)):
                ay  = str(ay_row[j])  if ay_row[j]  else ''
                hf  = str(hf_row[j])  if hf_row[j]  else ''
                tip = str(tip_row[j]) if tip_row[j] else ''
                if ay == ay_adi and str(son_hafta) == hf and tip == 'SDG':
                    all_col = j + 3
                    for i in range(7, min(50, len(rows))):
                        v = rows[i][all_col] if all_col < len(rows[i]) else None
                        if v and v != '#REF!':
                            try:
                                float(v); grc_all_col = all_col; break
                            except: pass
                    if grc_all_col: break

    # Hiç bulunamazsa son dolu SDG+3 sütunu
    if grc_all_col is None:
        for j in reversed(range(len(tip_row))):
            tip = str(tip_row[j]) if tip_row[j] else ''
            if tip == 'SDG':
                all_col = j + 3
                for i in range(7, min(50, len(rows))):
                    v = rows[i][all_col] if all_col < len(rows[i]) else None
                    if v and v != '#REF!':
                        try:
                            float(v); grc_all_col = all_col; break
                        except: pass
                if grc_all_col: break

    # AY Sezon Hedef % — ay bağımsız sütun
    for j in range(len(tip_row)):
        tip = str(tip_row[j]) if tip_row[j] else ''
        if 'AY Sezon Hedef' in tip:
            ay_sezon_col = j; break

    def extract(col):
        if col is None: return {}
        d = {}
        for i in range(7, min(110, len(rows))):
            row = rows[i]
            if not (row[0] and str(row[0]).strip()): continue
            key = str(row[0])
            if key in d: continue
            v = row[col] if col < len(row) else None
            if v and v != '#REF!':
                try: d[key] = float(v)
                except: pass
        return d

    grc = extract(grc_all_col)
    hdf = extract(hdf_all_col)
    ay_sezon = extract(ay_sezon_col)

    return {
        'gerceklesen_mtd': grc,
        'hedef_mtd':       hdf,
        'ay_sezon_hedef':  ay_sezon,
        'cat_labels':      {str(rows[i][0]): str(rows[i][1]) if len(rows[i]) > 1 and rows[i][1] else str(rows[i][0])
                            for i in range(7, min(110, len(rows))) if rows[i][0]},
        'kaynak': sheet_name,
        'kullanilan_ay': ay_adi,
        'kullanilan_hafta': son_hafta,
    }

def _empty_result():
    return {'gerceklesen_mtd':{}, 'hedef_mtd':{}, 'ay_sezon_hedef':{},
            'cat_labels':{}, 'kaynak':None, 'kullanilan_ay':None, 'kullanilan_hafta':None}


# ─────────────────────────────────────────────────────────────────────────────
# Rapor PPTX → Hedef / Gerçekleşen + KPI (OCR)
# ─────────────────────────────────────────────────────────────────────────────
def _ocr_lfl_row(img_rgb):
    import pytesseract
    arr = np.array(img_rgb)
    h, w = arr.shape[:2]
    lfl = img_rgb.crop((0, 28, w, min(130, h)))
    arr_c = np.array(lfl)
    out = np.ones_like(arr_c) * 255
    wm = (arr_c[:,:,0]>180)&(arr_c[:,:,1]>180)&(arr_c[:,:,2]>180)
    bm = (arr_c[:,:,0]<80) &(arr_c[:,:,1]<80) &(arr_c[:,:,2]<80)
    out[bm]=[0,0,0]; out[wm]=[0,0,0]
    big = Image.fromarray(out.astype(np.uint8)).resize(
        (out.shape[1]*5, out.shape[0]*5), Image.LANCZOS)
    return pytesseract.image_to_string(big, lang='eng', config='--psm 6 --oem 3')

def _normalize_pct(s):
    s = s.strip(); neg = s.startswith('-')
    core = s.lstrip('-').rstrip('%')
    if ',' not in core and '.' not in core and len(core)==3:
        core = core[:2]+','+core[2]
    return ('-' if neg else '')+core+'%'

def _parse_lfl_line(line):
    pcts_raw = re.findall(r'-?[\d]+[.,][\d]+%|-?[\d]{3,}%', line)
    pcts = [_normalize_pct(p) if re.match(r'-?[\d]{2,}%$',p) else p for p in pcts_raw]
    # Para: ₺/£/4 (OCR bozuk olabilir) + sayı (3 haneli veya nokta/virgülle)
    money = re.findall(r'(?:[£₺4#])\s*(\d{1,3}(?:[.,]\d{3})?)', line)
    if not money:
        money = re.findall(r'(?<!\d)(\d{1,2}\.\d{3})(?!\d)', line)
    floats = [f for f in re.findall(r'\b\d+[,\.]\d{2}\b', line)
              if float(f.replace(',','.')) < 10]
    result = {}
    if len(money)>=1: result['sepet_tl_ty']    = money[0].replace(',','.')
    if len(money)>=2: result['birim_fiyat_ty'] = money[1].replace(',','.')
    if floats:        result['upt_ty']         = floats[0]
    labels = ['net_tl_vs_ly','net_adet_vs_ly','islem_adet_vs_ly',
              'sepet_tl_vs_ly','upt_vs_ly','birim_fiyat_vs_ly']
    for i,lbl in enumerate(labels):
        if i < len(pcts): result[lbl] = pcts[i]
    return result

def parse_rapor(rapor_path):
    """Rapor pptx slayt 2 metin parse + slayt 3 KPI OCR."""
    with zipfile.ZipFile(rapor_path) as z:
        slide2_txt = re.findall(r'<a:t>([^<]+)</a:t>',
                                z.read('ppt/slides/slide2.xml').decode('utf-8'))
    full = ''.join(slide2_txt)
    result = {'hedef_mtl': None, 'grc_mtl': None, 'hgo': None, 'ay_etiket': ''}

    # Hedef
    m = re.search(r'hedefi?\s*([\d,.\s]+?)\s*MTL', full, re.I)
    if m:
        raw = m.group(1).strip().replace(' ','')
        if ',' in raw:
            after_comma = raw.split(',')[-1]
            if len(after_comma) == 3:
                raw = raw.replace(',', '')
            else:
                raw = raw.replace('.','').replace(',','.')
        else:
            raw = raw.replace('.','')
        try: result['hedef_mtl'] = float(raw)
        except: pass

    # Gerçekleşen delta
    m2 = re.search(r'hedefe göre\s*(-?[\d.,\s]+?)\s*MTL', full, re.I)
    if m2 and result['hedef_mtl']:
        raw2 = m2.group(1).strip()
        neg  = raw2.startswith('-')
        raw2 = raw2.lstrip('-').strip()
        raw2 = raw2.replace(',', '.')
        parts = raw2.split('.')
        if len(parts) > 2:
            raw2 = ''.join(parts[:-1]) + '.' + parts[-1]
        try:
            delta = float(raw2) * (-1 if neg else 1)
            result['grc_mtl'] = result['hedef_mtl'] + delta
            result['hgo'] = round(result['grc_mtl']/result['hedef_mtl']*100, 1)
        except: pass

    # Ay etiketi
    m_ay = re.search(
        r"(Ocak|Şubat|Mart|Nisan|Mayıs|Haziran|Temmuz|Ağustos|Eylül|Ekim|Kasım|Aralık)'?(\d{2})",
        full)
    if m_ay: result['ay_etiket'] = f"{m_ay.group(1)}'{m_ay.group(2)}"

    # KPI OCR (slayt 3)
    try:
        with zipfile.ZipFile(rapor_path) as z:
            rels3 = z.read('ppt/slides/_rels/slide3.xml.rels').decode()
            m_img = re.search(r'Target="\.\./media/([^"]+\.png)"', rels3)
            if m_img:
                img = Image.open(z.open(f'ppt/media/{m_img.group(1)}')).convert('RGB')
                lfl_txt = _ocr_lfl_row(img)
                for line in lfl_txt.split('\n'):
                    if '%' in line and ('.' in line or ',' in line):
                        kpi = _parse_lfl_line(line)
                        if kpi:
                            result.update(kpi)
                            break
    except Exception:
        pass

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Format yardımcıları
# ─────────────────────────────────────────────────────────────────────────────
def fmt_pct(v, decimals=1):
    if v is None: return "-"
    return f"%{v*100:.{decimals}f}".replace(".", ",")

def fmt_mtl(v):
    """3050.0 → '3.050 MTL'"""
    if v is None: return "-"
    return f"{round(v):,}".replace(",", ".") + " MTL"
